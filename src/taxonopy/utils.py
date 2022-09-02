# -*- coding: utf-8 -*-

import os
import tempfile
import importlib

import graphviz
from anytree import PreOrderIter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from .db import JSONDataBase, make_query
from .schema import (RecordBuilderBase,
                     SCHTree,
                     copy_node_to_record,
                     get_node_attr,
                     get_node_path,
                     record_has_node)


class FlatRecordBuilder(RecordBuilderBase):
    
    def __init__(self, schema, title_sep=":", value_sep=", "):
        super().__init__(schema)
        self._title_sep = title_sep
        self._value_sep = value_sep
    
    def build(self, existing, strict=False):
        
        self._iters = [iter([self._schema.root_node])]
        record = SCHTree()
        record = self._build(record, existing, strict=strict)
        
        return record
    
    def _build_node(self, record, node, existing, strict=False):
        
        node_attr = get_node_attr(node, blacklist=["name"])
        
        # See if node requires data first
        if "type" in node_attr:
            
            self._select_from_type(record,
                                   node,
                                   node_attr,
                                   existing,
                                   strict=strict)
            
            # If a typed node has no value and no children then we're done
            if "value" not in node_attr or not node.children:
                return
        
        # Now see if the node is a header for list selection
        if "inquire" in node_attr and node_attr["inquire"] == "list":
            self._select_from_list(record,
                                   node,
                                   node_attr,
                                   existing,
                                   strict=strict)
            return
        
        # Now see if the node is a header for checkbox selection
        if "inquire" in node_attr and node_attr["inquire"] == "checkbox":
            self._select_from_check(record,
                                    node,
                                    node_attr,
                                    existing,
                                    strict=strict)
            return
        
        copy_node_to_record(record, node, **node_attr)
        
        if len(node.children) < 1: return
        
        new_iter = PreOrderIter(node, maxlevel=2)
        next(new_iter)
        
        self._iters.append(new_iter)
    
    def _select_from_type(self, record,
                                node,
                                node_attr,
                                existing,
                                strict=False):
        
        if self._set_node_type(node, node_attr, existing, strict=strict):
            copy_node_to_record(record, node, **node_attr)
    
    def _set_node_type(self, node, node_attr, existing, strict=False):
        
        required = False
        node_title = self._get_node_title(node)
        existing_value = existing[node_title]
        
        if "required" in node_attr: required = bool(node_attr["required"])
        
        if required and existing_value is None:
            err_msg = (f"Node {node.name} is required, but existing record "
                        "has no value.")
            raise ValueError(err_msg)
        
        if existing_value is None: return False
        
        if "import" in node_attr:
            import_str = (f'{node_attr["import"]} = '
                      'importlib.import_module(node_attr["import"])')
            exec(import_str)
        
        val_type = eval(node_attr["type"])
        
        try:
            val_type(existing_value)
        except ValueError as e:
            if strict: raise ValueError(e)
            return False
        
        node_attr["value"] = existing_value
        
        return True
    
    def _select_from_list(self, record,
                                node,
                                node_attr,
                                existing,
                                strict=False):
        
        required = False
        node_path = get_node_path(node)
        node_title = self._get_node_title(node)
        existing_value = existing[node_title]
        
        # If not yet added, check if node is required
        if (record_has_node(record, node_path) or
            "required" in node_attr and node_attr["required"] == "True"):
            required = True
        
        # Gather names of children
        choices = [x.name for x in node.children]
        
        if required and existing_value is None:
            err_msg = (f"Node {node.name} is required, but existing record"
                        "has no value.")
            raise ValueError(err_msg)
        elif required and existing_value not in choices:
            err_msg = f"Entry for required node {node.name} is not valid"
            raise ValueError(err_msg)
        elif existing_value is None:
            return
        elif existing_value not in choices and strict:
            err_msg = (f"Value '{existing_value}' is not a valid choice for "
                       f"node {node.name}")
            raise ValueError(err_msg)
        elif existing_value not in choices:
            return
        
        # Add the node to the record if required
        copy_node_to_record(record, node, **node_attr)
        
        choice_path = f"{node_path}/{existing_value}"
        chosen_node = self._schema.find_by_path(choice_path)
        chosen_node_attr = get_node_attr(chosen_node, blacklist=["name"])
        
        if "type" in chosen_node_attr:
            self._select_from_type(record,
                                   chosen_node,
                                   chosen_node_attr,
                                   existing)
        
        copy_node_to_record(record, chosen_node, **chosen_node_attr)
        
        if len(chosen_node.children) < 1: return
        
        if "inquire" in chosen_node_attr:
            new_iter = PreOrderIter(chosen_node, maxlevel=1)
        else:
            new_iter = PreOrderIter(chosen_node, maxlevel=2)
            next(new_iter)
        
        self._iters.append(new_iter)
    
    def _select_from_check(self, record,
                                 node,
                                 node_attr,
                                 existing,
                                 strict=False):
        
        required = False
        node_path = get_node_path(node)
        node_title = self._get_node_title(node)
        existing_value = existing[node_title]
        
        # Check if required
        if "required" in node_attr and node_attr["required"] == "True":
            required = True
        
        if required and existing_value is None:
            err_msg = (f"Node {node.name} is required, but existing record"
                        "has no value.")
            raise ValueError(err_msg)
        
        if existing_value is None: return
        
        # Check existing values against choices
        choices = [x.name for x in node.children]
        existing_values = set(existing_value.split(self._value_sep))
        valid_values = list(set(choices) & existing_values)
        
        if required and not valid_values:
            
            err_msg = f"No valid entries found for required node {node.name}"
            raise ValueError(err_msg)
        
        elif strict and len(valid_values) != len(existing_values):
            
            bad_values = existing_values.difference(set(valid_values))
            bad_values_str = ", ".join(bad_values)
            
            if len(bad_values) == 1:
                noun = "choice"
            else:
                noun = "choices"
            
            err_msg = (f"Invalid {noun} '{bad_values_str}' given for node "
                       f"{node.name}")
            raise ValueError(err_msg)
        
        if not valid_values: return
        
        # Sort the values based on the order of choices
        sorter = {k:v for v,k in enumerate(choices)}
        valid_values.sort(key=sorter.get)
        
        # Add the parent node if it's not already in the tree
        copy_node_to_record(record, node, **node_attr)
        
        chosen_nodes = []
        
        for choice in valid_values:
            
            choise_path = f"{node_path}/{choice}"
            chosen_node = self._schema.find_by_path(choise_path)
            chosen_node_attr = get_node_attr(chosen_node,
                                             blacklist=["name"])
            
            if "type" in chosen_node_attr:
                self._select_from_type(record,
                                       chosen_node,
                                       chosen_node_attr,
                                       existing)
            
            if len(chosen_node.children) > 0:
                chosen_nodes.append(chosen_node)
            
            copy_node_to_record(record, chosen_node, **chosen_node_attr)
        
        if len(chosen_nodes) < 1: return
        
        new_iter = iter(chosen_nodes)
        self._iters.append(new_iter)
    
    def _get_node_title(self, node):
        
        if node.path == self._schema.root_node.path:
            return node.name
        
        return self._title_sep.join([str(x.name) for x in node.path
                                   if x.path != self._schema.root_node.path])


def get_root_value_ids(db):
    return {v.root_node.value: k for k, v in db.all().items()}


def dump_xl(out,
            schema,
            db,
            img_format='png',
            title_sep=":",
            value_sep=", "):
    
    # Remove xls or xlsx extension if added
    if out[-5:] == ".xlsx":
        out = out[:-5]
    elif out[-4:] == ".xls":
        out = out[:-4]
    
    # Add xlsx extension
    out += ".xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'DataBase'
    
    titles = _get_tree_titles(schema, sep=title_sep)
    required = _get_tree_attrs(schema.root_node, "required", False)
    ws.append(titles)
    
    query = make_query(titles[0])
    memdb = db.search(query)
    
    for record in memdb.to_records().values():
        
        row_values = [None] * len(titles)
        record_titles = _get_tree_titles(record, sep=title_sep)
        record_values = _get_tree_values(record.root_node,
                                         value_sep=value_sep)
        
        for col_title, col_value in zip(record_titles, record_values):
            col_idx = titles.index(col_title)
            row_values[col_idx] = col_value
        
        ws.append(row_values)
    
    for cell, black in zip(ws["1:1"], required):
        if black:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(bold=True, color="0070C0")
    
    for col in ws.columns:
        
        max_length = 0
        column = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = max_length * 1.1
        ws.column_dimensions[column].width = adjusted_width
    
    ws1 = wb.create_sheet("Schema")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        
        img_path = os.path.join(tmpdir, "temp") + "." + img_format
        render_tree(schema, img_path)
        
        img = Image(img_path)
        img.anchor = 'A1'
        
        scaling = 1200 / img.width
        img.width = 1200
        img.height *= scaling
        
        ws1.add_image(img)
        wb.save(out)


def load_xl(db_path,
            xl_path,
            schema,
            strict=False,
            progress=False,
            title_sep=":",
            value_sep=", "):
    
    schema_titles = _get_tree_titles(schema, sep=title_sep)
    builder = FlatRecordBuilder(schema, title_sep, value_sep)
    
    wb = load_workbook(xl_path)
    ws = wb['DataBase']
    titles = [cell.value for cell in ws[1] if cell.value is not None]
    
    if strict and not set(titles) <= set(schema_titles):
        
        extra_titles = set(titles) - set(schema_titles)
        extra_titles_str = ", ".join(extra_titles)
        
        if len(extra_titles) == 1:
            noun = "column"
        else:
            noun = "columns"
        
        err_msg = (f"Invalid {noun} '{extra_titles_str}' found")
        raise ValueError(err_msg)
    
    with JSONDataBase(db_path) as db:
        
        root_value_ids = get_root_value_ids(db)
        
        for values in ws.iter_rows(min_row=2, values_only=True):
            
            flat = {t: v for t, v in zip(titles, values)}
            if flat['Title'] is None: continue
            record = builder.build(flat, strict=strict)
            root_value = record.root_node.value
            
            if root_value in root_value_ids:
                doc_id = root_value_ids.pop(root_value)
                db.replace(doc_id, record)
                root_value_ids
            else:
                db.insert(record)
            
            if progress: print(".", end="", flush=True)
        
        db.remove(root_value_ids.values())
        
        if progress: print("\n", end="", flush=True)


def render_tree(tree, path):
    
    path, ext = os.path.splitext(path)
    img_format = ext[1:]
    
    dot = tree.to_dot()
    gv = graphviz.Source(dot)
    gv.format = img_format
    gv.render(path, cleanup=True)


def choice_count(path,
                 db,
                 schema):
    
    node = schema.find_by_path(path)
    
    if not hasattr(node, "inquire"): return
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") not in ["list", "checkbox"]): return
    
    count = {child.name: db.count(make_query(get_node_path(child)))
                                                 for child in node.children}
    
    missing_count = db.count(~make_query(path))
    if missing_count: count["None"] = missing_count
    
    return count


def find_non_matching_records(records_one,
                              records_two):
    
    records_one_name_dict = {tree.root_node.value: tree
                                       for tree in records_one.values()}
    records_two_name_dict = {tree.root_node.value: tree
                                       for tree in records_two.values()}
    
    records_one_name_set = set(records_one_name_dict.keys())
    records_two_name_set = set(records_two_name_dict.keys())
    
    matching = list(records_one_name_set & records_two_name_set)
    missing = list(records_one_name_set ^ records_two_name_set)
    
    for name in matching:
        if records_one_name_dict[name] == records_two_name_dict[name]:
            continue
        missing.append(name)
    
    return missing


def _get_tree_titles(tree, sep=":"):
    root_node = tree.root_node
    titles = [root_node.name]
    child_titles = _get_child_titles(root_node, sep=sep)
    titles.extend(child_titles)
    return titles


def _get_child_titles(node, parent=None, sep=":"):
    
    titles = []
    
    is_option = False
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") in ["list", "checkbox"]):
        is_option = True
    
    for child in node.children:
        
        if (is_option and
            not (child.children or hasattr(child, "type"))): continue
        
        if parent is None:
            name = child.name
        else:
            name = f"{parent}{sep}{child.name}"
        
        titles.append(name)
        
        if child.children:
            child_titles = _get_child_titles(child, name, sep)
            titles.extend(child_titles)
    
    return titles


def _get_tree_values(node, value_sep=", "):
    
    values = []
    prep_values = []
    
    if hasattr(node, "value"):
        prep_values.append(getattr(node, "value"))
    
    is_option = False
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") in ["list", "checkbox"]):
        is_option = True
    
    if is_option:
        for child in node.children:
            prep_values.append(child.name)
    
    value = value_sep.join(sorted(prep_values))
    values.append(value)
    
    for child in node.children:
        
        if (is_option and
            not (child.children or hasattr(child, "type"))): continue
                
        child_values = _get_tree_values(child, value_sep)
        values.extend(child_values)
    
    return values


def _get_tree_attrs(node, attr, default=None):
    
    results = []
    
    if hasattr(node, attr):
        results.append(getattr(node, attr))
    else:
        results.append(default)
    
    is_option = False
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") in ["list", "checkbox"]):
        is_option = True
    
    for child in node.children:
        
        if (is_option and
            not (child.children or hasattr(child, "type"))): continue
        
        child_results = _get_tree_attrs(child, attr, default)
        results.extend(child_results)
    
    return results
