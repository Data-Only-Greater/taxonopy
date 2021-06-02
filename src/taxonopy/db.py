# -*- coding: utf-8 -*-

import os
import sys
import tempfile
import itertools
from collections import OrderedDict
from collections.abc import Iterable

import graphviz
import inquirer
from anytree.resolver import ChildResolverError
from inquirer.render.console import ConsoleRender
from tinydb import table, TinyDB, Query
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from .tree import SCHTree
from .record import CLITheme, CLIRecordBuilder, FlatRecordBuilder


class DataBase:
    
    def __init__(self, db_path="db.json", check_existing=False):
        
        if check_existing and not os.path.isfile(db_path):
            raise IOError(f"Path {db_path} does not contain a valid database")
        
        self._db = TinyDB(db_path,
                          sort_keys=True,
                          indent=4,
                          separators=(',', ': '))
    
    def add(self, record):
        self._db.insert(record.to_dict())
    
    def all(self):
        sorter = _get_doc_sorter()
        return OrderedDict((doc.doc_id, SCHTree.from_dict(dict(doc)))
                                   for doc in sorted(self._db, key=sorter))
    
    def count(self, node_path, value=None, exact=False):
        return self._db.count(_make_query(node_path, value, exact))
    
    def get(self, node_path, value=None, exact=False):
        query = _make_query(node_path, value, exact)
        sorter = _get_doc_sorter()
        return OrderedDict((doc.doc_id, SCHTree.from_dict(dict(doc)))
                       for doc in sorted(self._db.search(query), key=sorter))
    
    def replace(self, doc_id, record):
        self._db.remove(doc_ids=[doc_id])
        self._db.insert(table.Document(record.to_dict(), doc_id=doc_id))


def new_record(schema_path="schema.json",
               db_path="db.json"):
    
    schema = SCHTree.from_json(schema_path)
    db = DataBase(db_path)
    
    builder = CLIRecordBuilder(schema)
    record = None
    
    while True:
        
        record = builder.build(record)
        node = record.root_node
        print(record)
        
        message = f"Store record with {node.name} '{node.value}'?"
        
        try:
            choice = inquirer.list_input(message,
                                         render=ConsoleRender(
                                                             theme=CLITheme()),
                                         choices=['yes', 'retry', 'quit'],
                                         default="yes")
        except KeyboardInterrupt:
            sys.exit()
        
        if choice == "quit": return
        if choice == "retry": continue
        
        count = db.count(f"{node.name}", node.value)
        
        if count > 0:
            
            message = (f"A record with {node.name} '{node.value}' already "
                        "exists. Add another?")
            
            try:
                choice = inquirer.list_input(
                                        message,
                                        render=ConsoleRender(
                                                             theme=CLITheme()),
                                        choices=['yes', 'retry', 'quit'],
                                        default="retry")
            except KeyboardInterrupt:
                sys.exit()
            
            if choice == "quit": return
            if choice == "retry": continue
        
        db.add(record)
        return


def show_count(path,
               value=None,
               exact=False,
               db_path="db.json"):
    db = DataBase(db_path, check_existing=True)
    count = db.count(path, value, exact)
    msg = f"{path}: {count}"
    print(msg)


def show_nodes(paths=None, db_path="db.json"):
    
    def _get_msg(node):
        
        msg_str = f"{node.name}"
        
        if hasattr(node, "value"):
            msg_str += f": {node.value}"
            return msg_str
            
        if hasattr(node, "inquire"):
            
            child_names = []
            sorter = _get_node_sorter()
            
            for child in sorted(node.children, key=sorter):
                child_names.append(child.name)
            
            child_str = ", ".join(child_names)
            
            if len(child_names) == 1:
                msg_str += f": {child_str}"
            elif len(child_names) > 1:
                msg_str += ": {" + child_str + "}"
            
            return msg_str
        
        return msg_str
    
    multi_field = False
    
    if not _is_iterable(paths):
        paths = (paths,)
    
    if len(paths) > 2: multi_field = True
    db = DataBase(db_path, check_existing=True)
    records = db.all()
    
    msg_rows = []
    
    for record in records.values():
        
        msgs = []
        
        for path in paths:
            
            if path is None:
                node = record.root_node
            else:
                try: 
                    node = record.find_by_path(path)
                except ChildResolverError:
                    msgs.append(False)
                    continue
            
            msgs.append(_get_msg(node))
        
        if not multi_field:
            if all(msgs): msg_rows.append(msgs)
            continue
        
        filtered_msgs = tuple(x for x in msgs if x)
        if len(filtered_msgs) < 2: continue
        msg_rows.append(filtered_msgs)
    
    widths = [len(max(msgs, key=len)) + 1
                  for msgs in itertools.zip_longest(*msg_rows, fillvalue="")]
    
    for x in msg_rows:
        
        final_str = ""
        
        for i, msg in enumerate(x):
            final_str += f'{msg: <{widths[i]}}'
            if i < len(widths) - 1: final_str += "| "
        
        print(final_str)


def show_records(path, value=None, exact=False, db_path="db.json"):
    db = DataBase(db_path, check_existing=True)
    for record in db.get(path, value, exact).values(): print(record)


def update_records(path,
                   value=None,
                   exact=False,
                   node_path=None,
                   schema_path="schema.json",
                   db_path="db.json"):
    
    schema = SCHTree.from_json(schema_path)
    builder = CLIRecordBuilder(schema)
    db = DataBase(db_path, check_existing=True)
    
    for doc_id, record in db.get(path, value, exact).items():
        
        node = record.root_node
        message = f"Update record with {node.name} '{node.value}'?"
        
        try:
            choice = inquirer.list_input(message,
                                         render=ConsoleRender(
                                                             theme=CLITheme()),
                                         choices=['yes', 'no', 'quit'],
                                         default="yes")
        except KeyboardInterrupt:
                sys.exit()
            
        if choice == "quit": return
        if choice == "no": continue
        
        while True:
            
            updated = builder.build(record, node_path)
            print(updated)
            
            message = "Store updated record?"
            
            try:
                choice = inquirer.list_input(
                                        message,
                                        render=ConsoleRender(theme=CLITheme()),
                                        choices=['yes',
                                                 'no',
                                                 'retry',
                                                 'quit'],
                                        default="yes")
            except KeyboardInterrupt:
                sys.exit()
            
            if choice == "quit": return
            if choice == "retry": continue
            if choice == "no": break
            
            db.replace(doc_id, updated)
            break


def dump_xl(out,
            schema_path="schema.json",
            db_path="db.json",
            img_format='png',
            title_sep=":",
            value_sep=", "):
    
    # Remove xls or xlsx extension if added
    if out[-5:] == ".xlsx":
        out = out[:-5]
    elif out[-4:] == ".xls":
        out = out[:-4]
    
    # Add xlsx extension
    out += ".xlsx"
    
    schema = SCHTree.from_json(schema_path)
    db = DataBase(db_path, check_existing=True)
    wb = Workbook()
    
    ws = wb.active
    ws.title = 'DataBase'
    
    titles = _get_tree_titles(schema, sep=title_sep)
    ws.append(titles)
    
    for record in db.get("Title").values():
        
        row_values = [None] * len(titles)
        record_titles = _get_tree_titles(record, sep=title_sep)
        record_values = _get_tree_values(record.root_node,
                                         title_sep=title_sep,
                                         value_sep=value_sep)
        
        for col_title, col_value in zip(record_titles, record_values):
            col_idx = titles.index(col_title)
            row_values[col_idx] = col_value
        
        ws.append(row_values)
    
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    
    for col in ws.columns:
        
        max_length = 0
        column = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = max_length * 1.1
        ws.column_dimensions[column].width = adjusted_width
    
    ws1 = wb.create_sheet("Schema")
    
    dot = schema.to_dot()
    gv = graphviz.Source(dot)
    gv.format = img_format
    
    with tempfile.TemporaryDirectory() as tmpdir:
        
        img_name = os.path.join(tmpdir, "temp")
        img_path = gv.render(img_name)
        img = Image(img_path)
        img.anchor = 'A1'
        
        ws1.add_image(img)
        wb.save(out)


def load_xl(db_path, xl_path, schema_path="schema.json", append=False):
    
    if not append and os.path.exists(db_path):
        os.remove(db_path)
    
    schema = SCHTree.from_json(schema_path)
    builder = FlatRecordBuilder(schema)
    
    wb = load_workbook(xl_path)
    db = DataBase(db_path)
    
    ws = wb['DataBase']
    titles = [cell.value for cell in ws[1]]
    
    for values in ws.iter_rows(min_row=2, values_only=True):
        flat = {t: v for t, v in zip(titles, values)}
        record = builder.build(flat)
        db.add(record)


def _make_query(path, value=None, exact=False):
    
    query = Query()
    path_resolution = path.strip('/').split('/')
    result = query['L0'].exists()
    
    for i, path in enumerate(path_resolution):
        result &= query[f'L{i}'].any(query.name == path)
    
    if value is not None:
        
        if exact:
            test = lambda value, search: search == value
        else:
            test = lambda value, search: search in value
            
        result &= query[f'L{len(path_resolution) - 1}'].any(
                                            query.value.test(test, value))
    
    return result


def _get_doc_sorter(path=None, case_insenstive=True):
    
    node_sorter = _get_node_sorter(case_insenstive)
    
    def sorter(doc):
        
        tree = SCHTree.from_dict(dict(doc))
        
        if path is None:
            node = tree.root_node
        else:
            node = tree.find_by_path(path)
        
        return node_sorter(node)
    
    return sorter


def _get_node_sorter(case_insenstive=True):
    
    def sorter(node):
        
        has_value = False
        if hasattr(node, "value"): has_value = True
        
        if case_insenstive:
            name = node.name.lower()
            if has_value: value = node.value.lower()
        else:
            name = node.name
            if has_value: value = node.value
        
        if has_value:
            result = (name, value)
        else:
            result = name
        
        return result
    
    return sorter


def _is_iterable(obj):
    
    result = False
    excluded_types = (str, dict)
    
    if isinstance(obj, Iterable) and not isinstance(obj, excluded_types):
        result = True
    
    return result


def _get_tree_titles(tree, sep=":"):
    root_node = tree.root_node
    titles = [root_node.name]
    child_titles = _get_child_titles(root_node, sep=sep)
    titles.extend(child_titles)
    return titles


def _get_child_titles(node, parent=None, sep=":"):
    
    titles = []
    
    is_option = False
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") in ["list", "checkbox"]):
        is_option = True
    
    for child in node.children:
        
        if (is_option and
            not (child.children or hasattr(child, "type"))): continue
        
        if parent is None:
            name = child.name
        else:
            name = f"{parent}{sep}{child.name}"
        
        titles.append(name)
        
        if child.children:
            child_titles = _get_child_titles(child, name, sep)
            titles.extend(child_titles)
    
    return titles


def _get_tree_values(node, parent=None, title_sep=":", value_sep=", "):
        
    def has_value(node):
        return hasattr(node, "value")
    
    def get_value(node):
        return getattr(node, "value")
    
    values = []
    prep_values = []
    
    if has_value(node):
        prep_values.append(get_value(node))
    
    is_option = False
    if (hasattr(node, "inquire") and
        getattr(node, "inquire") in ["list", "checkbox"]):
        is_option = True
    
    if is_option:
        for child in node.children:
            prep_values.append(child.name)
    
    value = value_sep.join(sorted(prep_values))
    values.append(value)
    
    for child in node.children:
    
        if (is_option and
            not (child.children or hasattr(child, "type"))): continue
        
        if parent is None:
            name = child.name
        else:
            name = f"{parent}{title_sep}{child.name}"
                
        child_values = _get_tree_values(child, name, title_sep, value_sep)
        values.extend(child_values)
    
    return values
